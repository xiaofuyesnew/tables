from openpyxl import Workbook

# 创建一个工作簿
wb = Workbook()

# 创建一个工作表
ws = wb.create_sheet(0)

# 切换到我们创建的表
wb.active = 1

# 更改工作表的名称
ws.title = '我的工作表'

# 给表格填一些数据
ws['A1'] = '小明'
ws['A2'] = '小红'
ws['B1'] = 59
ws['B2'] = 98

# 做点计算
ws['A3'] = '他们相差'
ws['B3'] = abs(ws['B1'].value - ws['B2'].value)

# 保存这个工作簿
wb.save('./第一个工作簿.xlsx')
