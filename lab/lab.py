from openpyxl import load_workbook
# from openpyxl.styles import colors, Font, Fill, NamedStyle
# from openpyxl.styles import PatternFill, Border, Side, Alignment

wb = load_workbook('./test.xlsx', data_only=True)

source_ws = wb['7.14']

target_ws = wb['Sheet2']

source_data = source_ws['C5':'C25']

print(source_data)

first = 0
first_cor = ''
second = 0
second_cor = ''
third = 0
third_cor = ''

for i in range(len(source_data)):
    for j in range(len(source_data[i])):
        # print(source_data[i][j].row)
        # print(get_column_letter(source_data[i][j].column))
        # print(source_data[i][j])
        item_value = int(source_data[i][j].value)
        if item_value > first:
            third = second
            third_cor = second_cor
            second = first
            second_cor = first_cor
            first = item_value
            first_cor = source_data[i][j].row
        elif item_value > second:
            third = second
            third_cor = second_cor
            second = item_value
            second_cor = source_data[i][j].row
        elif item_value > third:
            third = item_value
            third_cor = source_data[i][j].row

print(first)
print(second)
print(third)
print(source_ws['A' + str(first_cor)].value)
print(source_ws['A' + str(second_cor)].value)
print(source_ws['A' + str(third_cor)].value)


target_ws['F9'] = first
target_ws['F11'] = second
target_ws['F13'] = third
target_ws['E9'] = source_ws['A' + str(first_cor)].value
target_ws['E11'] = source_ws['A' + str(second_cor)].value
target_ws['E13'] = source_ws['A' + str(third_cor)].value

wb.save('./test.xlsx')

print(source_ws)
print(target_ws)